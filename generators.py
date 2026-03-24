import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

DAYS     = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
DAY_KEYS = ["sun", "mon", "tue", "wed", "thu", "fri", "sat"]

# Brand
BRAND_DARK = "FF1A3A5C"
BRAND_MID  = "FF2E6DA4"
BRAND_LITE = "FFD6E4F0"
WHITE      = "FFFFFFFF"
BLACK      = "FF000000"
GRAY_ROW   = "FFF5F5F5"
GRAY_ALT   = "FFEAEAEA"
ACCENT     = "FF0078D4"


def _thin_border():
    s = Side(style="thin", color=BLACK)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value="", bold=False, size=10, color=BLACK,
          bg=None, align="left", wrap=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if number_format:
        c.number_format = number_format
    return c


def generate_excel(emp_name: str, week: str, client: str, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    week_dt    = datetime.strptime(week, "%Y-%m-%d")
    week_end   = week_dt + timedelta(days=6)
    week_label = f"{week_dt.strftime('%B %d')} – {week_end.strftime('%B %d, %Y')}"

    col_widths = [28, 14, 7, 7, 7, 7, 7, 7, 7, 9, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    ws.merge_cells("A1:K1")
    hdr           = ws["A1"]
    hdr.value     = "AERIS TECHNICAL SOLUTIONS — TIMESHEET"
    hdr.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    hdr.fill      = PatternFill("solid", fgColor=BRAND_DARK)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 12):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=BRAND_DARK)

    for r, (lbl, val, lbl2, val2) in enumerate([
        ("Employee:", emp_name, "Week:", week_label),
        ("Client:",   client,   "Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ], start=2):
        _cell(ws, r, 1, lbl,  bold=True, bg=BRAND_LITE, color=BRAND_DARK)
        ws.merge_cells(f"B{r}:D{r}")
        _cell(ws, r, 2, val,  bg=BRAND_LITE, color=BRAND_DARK)
        _cell(ws, r, 5, lbl2, bold=True, bg=BRAND_LITE, color=BRAND_DARK)
        ws.merge_cells(f"F{r}:I{r}")
        _cell(ws, r, 6, val2, bg=BRAND_LITE, color=BRAND_DARK)

    # Header row
    ws.row_dimensions[5].height = 20
    for col, h in enumerate(["Pay Code", "PO Number"] + DAYS + ["Total", "Notes"], 1):
        c = _cell(ws, 5, col, h, bold=True, color=WHITE, bg=BRAND_DARK, align="center", size=9)
        c.border = _thin_border()

    if not rows:
        rows = [{}]

    row_num = 6
    for i, row in enumerate(rows):
        bg = GRAY_ROW if i % 2 == 0 else GRAY_ALT
        ws.row_dimensions[row_num].height = 18
        _cell(ws, row_num, 1, row.get("description", "Regular"), bg=bg).border = _thin_border()
        _cell(ws, row_num, 2, row.get("po_number", ""), bg=bg, align="center").border = _thin_border()

        for d, key in enumerate(DAY_KEYS):
            c = ws.cell(row=row_num, column=3 + d)
            c.value = row.get("days", {}).get(key)
            c.font  = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill  = PatternFill("solid", fgColor=bg)
            c.border = _thin_border()
            c.number_format = "0.##"

        day_range = f"C{row_num}:I{row_num}"
        tc = ws.cell(row=row_num, column=10)
        tc.value = f"=SUM({day_range})"
        tc.font  = Font(name="Arial", bold=True, size=10, color=ACCENT)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.fill  = PatternFill("solid", fgColor=bg)
        tc.border = _thin_border()
        tc.number_format = "0.##"

        _cell(ws, row_num, 11, row.get("notes", ""), bg=bg, wrap=True, size=9).border = _thin_border()
        row_num += 1

    # Totals row
    ws.row_dimensions[row_num].height = 20
    _cell(ws, row_num, 1, "WEEKLY TOTALS", bold=True, color=WHITE, bg=BRAND_MID, align="right").border = _thin_border()
    ws.cell(row=row_num, column=2).fill   = PatternFill("solid", fgColor=BRAND_MID)
    ws.cell(row=row_num, column=2).border = _thin_border()

    for col in range(3, 11):
        cl = get_column_letter(col)
        tc = ws.cell(row=row_num, column=col)
        tc.value = f"=SUM({cl}6:{cl}{row_num - 1})"
        tc.font  = Font(name="Arial", bold=True, size=10, color=WHITE)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.fill  = PatternFill("solid", fgColor=BRAND_MID)
        tc.border = _thin_border()
        tc.number_format = "0.##"

    ws.cell(row=row_num, column=11).fill   = PatternFill("solid", fgColor=BRAND_MID)
    ws.cell(row=row_num, column=11).border = _thin_border()

    ws.freeze_panes = "C6"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_pdf(emp_name: str, week: str, client: str, rows: list) -> bytes:
    buf      = io.BytesIO()
    week_dt  = datetime.strptime(week, "%Y-%m-%d")
    week_end = week_dt + timedelta(days=6)
    wlabel   = f"{week_dt.strftime('%B %d')} – {week_end.strftime('%B %d, %Y')}"

    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)

    BD    = colors.HexColor("#1A3A5C")
    BM    = colors.HexColor("#2E6DA4")
    BL    = colors.HexColor("#D6E4F0")
    GACC  = colors.HexColor("#0078D4")
    GR    = colors.HexColor("#F5F5F5")
    GA    = colors.HexColor("#EAEAEA")

    def ps(name, **kw):
        return ParagraphStyle(name, fontName="Helvetica", fontSize=9,
                              textColor=colors.black, **kw)

    story = []

    title_tbl = Table([[Paragraph("AERIS TECHNICAL SOLUTIONS — TIMESHEET",
        ps("t", fontName="Helvetica-Bold", fontSize=16, textColor=colors.white,
           alignment=1))]], colWidths=[10*inch])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), BD),
        ("TOPPADDING", (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 10),
    ]))
    story.append(title_tbl)

    meta = Table([[
        Paragraph("Employee:", ps("ml", fontName="Helvetica-Bold", textColor=BD)),
        Paragraph(emp_name,    ps("mv")),
        Paragraph("Week:",     ps("ml", fontName="Helvetica-Bold", textColor=BD)),
        Paragraph(wlabel,      ps("mv")),
        Paragraph("Client:",   ps("ml", fontName="Helvetica-Bold", textColor=BD)),
        Paragraph(client,      ps("mv")),
    ]], colWidths=[0.8*inch,2*inch,0.6*inch,2.4*inch,0.6*inch,3.6*inch])
    meta.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), BL),
        ("TOPPADDING",(0,0),(-1,-1), 5),("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING",(0,0),(-1,-1), 6),
        ("BOX",(0,0),(-1,-1), 0.5, BM),
    ]))
    story.append(meta)
    story.append(Spacer(1, 0.15*inch))

    ch  = ps("ch", fontName="Helvetica-Bold", textColor=colors.white, alignment=1)
    cv  = ps("cv", alignment=1)
    dl  = ps("dl", alignment=0)
    tot = ps("tot", fontName="Helvetica-Bold", textColor=GACC, alignment=1)
    wh  = ps("wh", fontName="Helvetica-Bold", textColor=colors.white, alignment=2)
    wv  = ps("wv", fontName="Helvetica-Bold", textColor=colors.white, alignment=1)

    hrow = ([Paragraph("Pay Code", ch), Paragraph("PO Number", ch)]
            + [Paragraph(d, ch) for d in DAYS]
            + [Paragraph("Total", ch), Paragraph("Notes", ch)])

    if not rows:
        rows = [{}]

    drows = []
    for row in rows:
        dvals = [row.get("days", {}).get(k) for k in DAY_KEYS]
        rt    = sum(float(v) for v in dvals if v is not None)
        drows.append(
            [Paragraph(row.get("description","Regular"), dl),
             Paragraph(row.get("po_number",""), cv)]
            + [Paragraph(str(v) if v is not None else "", cv) for v in dvals]
            + [Paragraph(f"{rt:.2f}", tot), Paragraph(row.get("notes",""), dl)]
        )

    day_tots = [sum(float(r.get("days",{}).get(k) or 0) for r in rows) for k in DAY_KEYS]
    grand    = sum(day_tots)
    frow     = ([Paragraph("WEEKLY TOTALS", wh), Paragraph("", wv)]
                + [Paragraph(f"{v:.2f}" if v else "", wv) for v in day_tots]
                + [Paragraph(f"{grand:.2f}", wv), Paragraph("", wv)])

    cw   = [2.4*inch,1.1*inch]+[0.58*inch]*7+[0.7*inch,1.7*inch]
    tbl  = Table([hrow]+drows+[frow], colWidths=cw, repeatRows=1)
    cmds = [
        ("BACKGROUND",(0,0),(-1,0), BD),
        ("BACKGROUND",(0,-1),(-1,-1), BM),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#BBBBBB")),
        ("BOX",(0,0),(-1,-1),0.8,BD),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]
    for i in range(1, len(drows)+1):
        cmds.append(("BACKGROUND",(0,i),(-1,i), GR if i%2==1 else GA))
    tbl.setStyle(TableStyle(cmds))
    story.append(tbl)

    story.append(Spacer(1, 0.35*inch))
    sig = ps("sig", textColor=colors.HexColor("#666666"))
    sig_tbl = Table([[
        Paragraph("Employee Signature: _________________________________", sig),
        Paragraph("", sig),
        Paragraph("Date: ____________________", sig),
    ]], colWidths=[4*inch,2*inch,4*inch])
    sig_tbl.setStyle(TableStyle([
        ("TOPPADDING",(0,0),(-1,-1),8),
        ("LINEBELOW",(0,0),(0,0),0.5,colors.HexColor("#AAAAAA")),
        ("LINEBELOW",(2,0),(2,0),0.5,colors.HexColor("#AAAAAA")),
    ]))
    story.append(sig_tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()
